# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as num_styles)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = '/sessions/ecstatic-sweet-hamilton/mnt/outputs/NI_Staff_Cost_Planner.xlsx'

wb = Workbook()

# ─────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────
def font(bold=False, color='000000', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def border(style='thin'):
    s = Side(style=style, color='C0C0C0')
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s = Side(style='thin', color='C0C0C0')
    return Border(bottom=s)

NAVY     = '1A3C5E'
TEAL     = '1A5E4E'
LTBLUE   = 'E8F2FB'
LTGREEN  = 'E8F5EE'
LTYELLOW = 'FFF9E6'
LTRED    = 'FDE8E8'
GRAY     = 'F4F7FA'
MIDGRAY  = 'DDE4EA'
WHITE    = 'FFFFFF'
INPUTBLUE = '0000FF'
FORMULABLK = '000000'
LINKGREEN  = '008000'
INPUTYELLOW = 'FFFF00'

FMT_USD  = '$#,##0;($#,##0);"-"'
FMT_PCT  = '0.0%;(0.0%);"-"'
FMT_NUM  = '#,##0;(#,##0);"-"'

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, size=10, span=1, align='left'):
    cell = ws.cell(row, col, text)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return cell

def inp(ws, row, col, value, fmt=None, comment=None, yellow=False):
    cell = ws.cell(row, col, value)
    cell.font = font(color=INPUTBLUE, bold=False)
    cell.fill = fill(INPUTYELLOW if yellow else 'FAFCFF')
    cell.border = border()
    cell.alignment = Alignment(horizontal='right', vertical='center')
    if fmt: cell.number_format = fmt
    return cell

def formula(ws, row, col, expr, fmt=None, green=False):
    cell = ws.cell(row, col, expr)
    cell.font = font(color=LINKGREEN if green else FORMULABLK)
    cell.border = border()
    cell.alignment = Alignment(horizontal='right', vertical='center')
    if fmt: cell.number_format = fmt
    return cell

def lbl(ws, row, col, text, bold=False, indent=0, bg=None, align='left'):
    cell = ws.cell(row, col, (' ' * indent) + text)
    cell.font = font(bold=bold)
    cell.border = border()
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bg: cell.fill = fill(bg)
    return cell

def subtotal_row(ws, row, start_col, end_col, bg=GRAY):
    for c in range(start_col, end_col+1):
        ws.cell(row, c).fill = fill(bg)
        ws.cell(row, c).font = font(bold=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell):
    ws.freeze_panes = cell

# ─────────────────────────────────────────────────────────────
# SHEET 1: INSTRUCTIONS
# ─────────────────────────────────────────────────────────────
ws_inst = wb.active
ws_inst.title = 'Instructions'
ws_inst.sheet_properties.tabColor = NAVY

def inst_row(ws, r, text, bold=False, indent=0, bg=WHITE):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws.cell(r, 1, (' '*indent)+text)
    c.font = font(bold=bold, size=10)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 18

ws_inst.merge_cells('A1:G1')
t = ws_inst['A1']
t.value = 'NUTRITION INTERNATIONAL — STAFF COST SCENARIO PLANNER'
t.font = Font(name='Arial', bold=True, size=14, color=WHITE)
t.fill = fill(NAVY)
t.alignment = Alignment(horizontal='center', vertical='center')
ws_inst.row_dimensions[1].height = 36

ws_inst.merge_cells('A2:G2')
s = ws_inst['A2']
s.value = 'Corporate Finance & Business Planning Tool  ·  FY 2025–26'
s.font = Font(name='Arial', size=10, color='AAAAAA', italic=True)
s.fill = fill(NAVY)
s.alignment = Alignment(horizontal='center', vertical='center')
ws_inst.row_dimensions[2].height = 22

inst_rows = [
    (3, '', False, 0, WHITE),
    (4, 'PURPOSE', True, 0, LTBLUE),
    (5, 'This tool helps managers scenario-plan the true cost of hiring staff across all Nutrition International office locations.', False, 2, WHITE),
    (6, 'It includes three cost calculators: total annual employment cost, cost of replacing a departing employee, and cost of staff leave.', False, 2, WHITE),
    (7, '', False, 0, WHITE),
    (8, 'WORKBOOK TABS', True, 0, LTBLUE),
    (9, '1.  Country Data        — Employer contribution rates and leave entitlements for all 11 NI locations', False, 2, WHITE),
    (10, '2.  Assumptions         — Shared inputs used across all calculators (edit these cells to model scenarios)', False, 2, WHITE),
    (11, '3.  Employment Cost     — Annual cost of employing one person: salary + mandatory contributions + benefits', False, 2, WHITE),
    (12, '4.  Replacement Cost    — Full cost of replacing a departing employee (recruitment, onboarding, ramp-up)', False, 2, WHITE),
    (13, '5.  Leave Cost          — Cost when an employee goes on maternity, medical, or extended leave', False, 2, WHITE),
    (14, '6.  Compare Locations   — Side-by-side comparison of all three costs across selected NI locations', False, 2, WHITE),
    (15, '', False, 0, WHITE),
    (16, 'HOW TO USE', True, 0, LTBLUE),
    (17, 'Step 1:  Go to the Assumptions tab and enter the position details (salary, seniority, contract type).', False, 2, WHITE),
    (18, 'Step 2:  Adjust the recruitment and leave assumptions as needed for your specific scenario.', False, 2, WHITE),
    (19, 'Step 3:  Open Employment Cost, Replacement Cost, or Leave Cost tabs to see calculations for one location.', False, 2, WHITE),
    (20, 'Step 4:  Use the Compare Locations tab to view up to 4 locations side by side.', False, 2, WHITE),
    (21, '', False, 0, WHITE),
    (22, 'COLOUR CODING', True, 0, LTBLUE),
    (23, '  Blue text       = Input cells — change these to model different scenarios', False, 2, WHITE),
    (24, '  Black text      = Calculated formulas — do not edit', False, 2, WHITE),
    (25, '  Green text      = Cross-sheet links pulling from Country Data or Assumptions', False, 2, WHITE),
    (26, '  Yellow fill     = Key assumption cells that drive the model', False, 2, WHITE),
    (27, '', False, 0, WHITE),
    (28, 'DATA SOURCES & DISCLAIMERS', True, 0, LTBLUE),
    (29, 'Employer contribution rates sourced from: PwC Worldwide Tax Summaries (2025), ISSA Country Profiles, national statutory authorities.', False, 2, WHITE),
    (30, 'Exchange rates are indicative only. All figures are estimates — verify with local HR/Finance before budget submission.', False, 2, WHITE),
    (31, 'Leave entitlements reflect statutory minimums; NI policies may exceed these.', False, 2, WHITE),
    (32, '', False, 0, WHITE),
]
for r, text, bold, indent, bg in inst_rows:
    inst_row(ws_inst, r, text, bold, indent, bg)

set_col_widths(ws_inst, {'A':12,'B':12,'C':12,'D':12,'E':12,'F':12,'G':20})

# ─────────────────────────────────────────────────────────────
# SHEET 2: COUNTRY DATA
# ─────────────────────────────────────────────────────────────
ws_cd = wb.create_sheet('Country Data')
ws_cd.sheet_properties.tabColor = '2D6A4F'

# Title
ws_cd.merge_cells('A1:N1')
h = ws_cd['A1']
h.value = 'COUNTRY DATA — Employer Contribution Rates & Leave Entitlements (All NI Office Locations)'
h.font = Font(name='Arial', bold=True, size=11, color=WHITE)
h.fill = fill(TEAL)
h.alignment = Alignment(horizontal='center', vertical='center')
ws_cd.row_dimensions[1].height = 28

# Column headers (row 2)
cd_headers = [
    'Country', 'Flag', 'Region', 'City', 'Currency', 'FX Rate\n(1 USD =)',
    'Pension /\nRetirement\n(% salary)',
    'Health /\nSocial\n(% salary)',
    'Other\nMandatory\n(% salary)',
    'TOTAL\nMandatory\n(%)',
    'Typical\nBenefits\n(% salary)',
    'Govt Leave\nPay Subsidy\n(%)',
    'Maternity\nLeave\n(weeks)',
    'Annual Leave\n(days/yr)'
]
for ci, h_text in enumerate(cd_headers, 1):
    c = ws_cd.cell(2, ci, h_text)
    c.font = Font(name='Arial', bold=True, size=9, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border()
ws_cd.row_dimensions[2].height = 48

# Data rows
countries = [
    # country, flag, region, city, currency, fx(units per 1 USD),
    # pension%, health%, other%, benefits%, govtLeave%, maternity_wks, annual_days
    ('Canada',      '🇨🇦', 'Americas (HQ)',       'Ottawa',        'CAD',  1/0.74,    5.95,  2.32,  0.0,   7.0,  55.0, 17,  15),
    ('Kenya',       '🇰🇪', 'Africa',              'Nairobi',       'KES',  1/0.0077,  2.0,   6.25,  0.0,   4.0,   0.0, 13,  21),
    ('Ethiopia',    '🇪🇹', 'Africa',              'Addis Ababa',   'ETB',  1/0.0088,  11.0,  0.0,   0.0,   3.0,   0.0, 13,  16),
    ('Nigeria',     '🇳🇬', 'Africa',              'Abuja',         'NGN',  1/0.00065, 10.0,  0.0,   2.0,   4.5,   0.0, 12,  21),
    ('Senegal',     '🇸🇳', 'Africa (Sahel)',      'Dakar',         'XOF',  1/0.00163, 14.0,  7.0,   12.0,  3.0,   0.0, 14,  24),
    ('Tanzania',    '🇹🇿', 'Africa',              'Dar es Salaam', 'TZS',  1/0.00039, 10.0,  0.0,   4.5,   3.5,   0.0, 13,  28),
    ('India',       '🇮🇳', 'Asia (Regional HQ)', 'New Delhi',     'INR',  1/0.012,   12.0,  3.25,  4.81,  3.5,   0.0, 26,  15),
    ('Bangladesh',  '🇧🇩', 'Asia',               'Dhaka',         'BDT',  1/0.0091,  10.0,  0.0,   4.17,  2.5,   0.0, 16,  10),
    ('Indonesia',   '🇮🇩', 'Asia',               'Jakarta',       'IDR',  1/0.000062,3.7,   4.0,   3.1,   3.0,   0.0, 13,  12),
    ('Pakistan',    '🇵🇰', 'Asia',               'Islamabad',     'PKR',  1/0.0036,  5.0,   6.0,   0.0,   3.0,   0.0, 12,  14),
    ('Philippines', '🇵🇭', 'Asia',               'Manila',        'PHP',  1/0.018,   9.5,   2.5,   2.0,   3.5,  80.0, 11,   5),
]

for ri, row in enumerate(countries, 3):
    country, flag, region, city, curr, fx, pension, health, other, benefits, govt_leave, mat_wks, ann_days = row
    total_mand = pension + health + other
    bg = LTBLUE if ri % 2 == 1 else WHITE

    vals = [country, flag, region, city, curr,
            round(fx, 1), pension, health, other, total_mand,
            benefits, govt_leave, mat_wks, ann_days]
    fmts = [None, None, None, None, None,
            '#,##0', FMT_PCT, FMT_PCT, FMT_PCT, FMT_PCT,
            FMT_PCT, FMT_PCT, '0', '0']

    for ci, (val, fmt_str) in enumerate(zip(vals, fmts), 1):
        cell = ws_cd.cell(ri, ci, val)
        cell.border = border()
        cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                   vertical='center')
        cell.fill = fill(bg)
        cell.font = font(size=10)
        if fmt_str: cell.number_format = fmt_str

        # Total mandatory column bold
        if ci == 10:
            cell.font = font(bold=True, size=10)
            cell.fill = fill(LTGREEN)

ws_cd.row_dimensions[1].height = 28
for r in range(3, 14):
    ws_cd.row_dimensions[r].height = 20

# Source note
ws_cd.merge_cells('A14:N14')
src = ws_cd['A14']
src.value = 'Sources: PwC Worldwide Tax Summaries (2025) · ISSA Country Profiles · National statutory authorities. Rates are indicative — verify before submission.'
src.font = Font(name='Arial', size=8, color='888888', italic=True)
src.alignment = Alignment(horizontal='left', vertical='center')

set_col_widths(ws_cd, {
    'A':14,'B':5,'C':18,'D':15,'E':8,'F':12,
    'G':11,'H':11,'I':11,'J':11,'K':11,'L':11,'M':10,'N':10
})

# ─────────────────────────────────────────────────────────────
# SHEET 3: ASSUMPTIONS
# ─────────────────────────────────────────────────────────────
ws_asmp = wb.create_sheet('Assumptions')
ws_asmp.sheet_properties.tabColor = 'F0A500'

ws_asmp.merge_cells('A1:E1')
t = ws_asmp['A1']
t.value = 'SCENARIO ASSUMPTIONS — Edit blue/yellow cells to model different scenarios'
t.font = Font(name='Arial', bold=True, size=11, color=WHITE)
t.fill = fill(NAVY)
t.alignment = Alignment(horizontal='center', vertical='center')
ws_asmp.row_dimensions[1].height = 28

def asmp_section(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row, 1, title)
    c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 20

def asmp_row(ws, row, label, value, unit, fmt_str=None, yellow=False, note=''):
    ws.row_dimensions[row].height = 18
    lc = ws.cell(row, 1, label)
    lc.font = font(size=10)
    lc.border = border()
    lc.alignment = Alignment(horizontal='left', vertical='center')

    vc = ws.cell(row, 2, value)
    vc.font = font(color=INPUTBLUE, size=10)
    vc.fill = fill(INPUTYELLOW if yellow else 'FAFCFF')
    vc.border = border()
    vc.alignment = Alignment(horizontal='right', vertical='center')
    if fmt_str: vc.number_format = fmt_str

    uc = ws.cell(row, 3, unit)
    uc.font = font(color='888888', size=9, italic=True)
    uc.border = border()
    uc.alignment = Alignment(horizontal='left', vertical='center')

    nc = ws.cell(row, 4, note)
    nc.font = font(color='888888', size=9, italic=True)
    nc.border = border()
    nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    return vc

# Section A: Position
asmp_section(ws_asmp, 2, 'A.  POSITION DETAILS')
asmp_row(ws_asmp, 3, 'Job Title', 'Program Officer', 'text', note='For reference only')
asmp_row(ws_asmp, 4, 'Seniority Level', 'Mid-Level (3-7 yrs)', 'text', note='Junior / Mid-Level / Senior / Director')
# Annual salary — KEY INPUT
sal_cell = asmp_row(ws_asmp, 5, 'Annual Base Salary', 45000, 'USD equivalent', FMT_USD, yellow=True,
                    note='Enter in USD. FX conversion shown on Country Data tab.')
sal_cell.fill = fill(INPUTYELLOW)
asmp_row(ws_asmp, 6, 'Contract Type', 'Permanent', 'text', note='Permanent / Fixed-Term / Consultant')

# Section B: Recruitment
asmp_section(ws_asmp, 8, 'B.  RECRUITMENT ASSUMPTIONS')
asmp_row(ws_asmp, 9, 'Recruitment Agency Fee', 0.18, '% of annual salary', FMT_PCT, yellow=True,
         note='Typical range: 12–25%. Use 0% for direct/internal recruitment.')
asmp_row(ws_asmp, 10, 'Job Advertising Cost', 400, 'USD', FMT_USD,
         note='Job boards, LinkedIn, local postings')
asmp_row(ws_asmp, 11, 'Total Interview Hours (panel)', 18, 'hours',
         note='Total interviewer-hours across all rounds')
asmp_row(ws_asmp, 12, 'Onboarding & Training Cost', 2000, 'USD', FMT_USD,
         note='Materials, induction time, IT setup')
asmp_row(ws_asmp, 13, 'Ramp-up Period', 3, 'months',
         note='Time until new hire reaches full productivity')
asmp_row(ws_asmp, 14, 'Productivity During Ramp-up', 0.55, '% of full output', FMT_PCT,
         note='Typical range: 40–70%')
asmp_row(ws_asmp, 15, 'Knowledge & Handover Loss', 0.10, '% of annual salary', FMT_PCT,
         note='Institutional knowledge, client relationships, documentation')

# Section C: Leave
asmp_section(ws_asmp, 17, 'C.  LEAVE ASSUMPTIONS')
asmp_row(ws_asmp, 18, 'Temp / Backfill Daily Rate', 200, 'USD/day', FMT_USD,
         note='Daily cost of temporary cover or backfill person')
asmp_row(ws_asmp, 19, 'Maternity Leave Duration (scenario)', 17, 'weeks',
         note='Override statutory minimum if NI policy is more generous')
asmp_row(ws_asmp, 20, 'Medical Leave Duration (scenario)', 4, 'weeks',
         note='Duration for the leave cost scenario')
asmp_row(ws_asmp, 21, 'Employer Continues Full Salary?', 'Yes', 'Yes/Partial/No',
         note='Yes = employer bears full salary; No = statutory benefit only')
asmp_row(ws_asmp, 22, 'Team Productivity Impact', 0.12, '% of leave salary', FMT_PCT,
         note='Overhead from covering duties, knowledge gaps, re-training')

# Named ranges for easy reference in formulas
# We'll use direct cell references in formulas instead

set_col_widths(ws_asmp, {'A':30,'B':14,'C':18,'D':28,'E':5})
ws_asmp.column_dimensions['E'].width = 5

# ─────────────────────────────────────────────────────────────
# SHEET 4: EMPLOYMENT COST
# ─────────────────────────────────────────────────────────────
ws_emp = wb.create_sheet('Employment Cost')
ws_emp.sheet_properties.tabColor = '1A3C5E'

# Title
ws_emp.merge_cells('A1:F1')
t = ws_emp['A1']
t.value = 'ANNUAL EMPLOYMENT COST CALCULATOR'
t.font = Font(name='Arial', bold=True, size=12, color=WHITE)
t.fill = fill(NAVY)
t.alignment = Alignment(horizontal='center', vertical='center')
ws_emp.row_dimensions[1].height = 30

# Location selector
ws_emp.merge_cells('A2:B2')
ws_emp['A2'] = 'Select Location:'
ws_emp['A2'].font = font(bold=True, size=10)
ws_emp['A2'].alignment = Alignment(horizontal='right', vertical='center')
ws_emp['A2'].fill = fill(LTBLUE)
ws_emp['A2'].border = border()

loc_cell = ws_emp.cell(2, 3, 'Canada')
loc_cell.font = font(color=INPUTBLUE, bold=True, size=11)
loc_cell.fill = fill(INPUTYELLOW)
loc_cell.border = border()
loc_cell.alignment = Alignment(horizontal='left', vertical='center')

# Dropdown validation
dv = DataValidation(
    type='list',
    formula1='"Canada,Kenya,Ethiopia,Nigeria,Senegal,Tanzania,India,Bangladesh,Indonesia,Pakistan,Philippines"',
    allow_blank=False, showDropDown=False
)
ws_emp.add_data_validation(dv)
dv.add('C2')

ws_emp.row_dimensions[2].height = 22

# Salary reference note
ws_emp.merge_cells('D2:F2')
n = ws_emp['D2']
n.value = '← Select country. Salary & assumptions from the Assumptions tab.'
n.font = font(color='888888', size=9, italic=True)
n.alignment = Alignment(horizontal='left', vertical='center')

# Helper: VLOOKUP column indices in Country Data (1-based)
# A=Country(1), J=TotalMandatory(10), G=Pension(7), H=Health(8), I=Other(9), K=Benefits(11), L=GovtLeave(12), M=Maternity(13), N=Annual(14)
CD = "'Country Data'!$A$3:$N$13"

# Section: Position Details
ws_emp.merge_cells('A4:F4')
ws_emp['A4'] = 'POSITION DETAILS'
ws_emp['A4'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_emp['A4'].fill = fill(TEAL)
ws_emp['A4'].alignment = Alignment(horizontal='left', vertical='center')

details = [
    (5, 'Job Title',             f"=Assumptions!B3"),
    (6, 'Seniority',             f"=Assumptions!B4"),
    (7, 'Annual Base Salary',    f"=Assumptions!B5"),
    (8, 'Contract Type',         f"=Assumptions!B6"),
    (9, 'Selected Country',      "=C2"),
    (10, 'Country Region',       f"=IFERROR(VLOOKUP(C2,{CD},3,0),\"-\")"),
    (11, 'Office City',          f"=IFERROR(VLOOKUP(C2,{CD},4,0),\"-\")"),
    (12, 'Local Currency',       f"=IFERROR(VLOOKUP(C2,{CD},5,0),\"-\")"),
    (13, 'Indicative FX Rate',   f"=IFERROR(VLOOKUP(C2,{CD},6,0),\"-\")"),
]

for row, label, expr in details:
    lbl(ws_emp, row, 1, label)
    ws_emp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = formula(ws_emp, row, 3, expr, green=True)
    if row == 7: c.number_format = FMT_USD
    if row == 13: c.number_format = '#,##0.0'
    ws_emp.row_dimensions[row].height = 18

# Section: Cost Breakdown
ws_emp.merge_cells('A15:F15')
ws_emp['A15'] = 'ANNUAL COST BREAKDOWN'
ws_emp['A15'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_emp['A15'].fill = fill(NAVY)
ws_emp['A15'].alignment = Alignment(horizontal='left', vertical='center')
ws_emp.row_dimensions[15].height = 20

cost_rows = [
    (16, 'Base Annual Salary',                    f"=Assumptions!B5",                   FMT_USD),
    (17, 'Mandatory Contributions (% of salary)', f"=IFERROR(VLOOKUP(C2,{CD},10,0),0)", FMT_PCT),
    (18, 'Mandatory Contributions (USD)',          f"=C16*C17",                          FMT_USD),
    (19, 'Employer Benefits % (est.)',             f"=IFERROR(VLOOKUP(C2,{CD},11,0),0)", FMT_PCT),
    (20, 'Employer Benefits (USD)',                f"=C16*C19",                          FMT_USD),
    (21, 'Total Employer On-Cost Burden %',        f"=(C18+C20)/C16",                    FMT_PCT),
]

for row, label, expr, fmt_str in cost_rows:
    lbl(ws_emp, row, 1, label)
    ws_emp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    formula(ws_emp, row, 3, expr, fmt=fmt_str, green=(row in [16,17,19]))
    ws_emp.row_dimensions[row].height = 18

# Total line
ws_emp.row_dimensions[22].height = 22
lbl(ws_emp, 22, 1, 'TOTAL ANNUAL EMPLOYMENT COST (USD)', bold=True)
ws_emp.merge_cells('A22:B22')
c_total = formula(ws_emp, 22, 3, "=C16+C18+C20", fmt=FMT_USD)
c_total.font = Font(name='Arial', bold=True, size=12, color=WHITE)
c_total.fill = fill(NAVY)
for col in range(1, 7):
    ws_emp.cell(22, col).fill = fill(NAVY)
    ws_emp.cell(22, col).font = Font(name='Arial', bold=True, color=WHITE, size=11)

# Contribution detail section
ws_emp.merge_cells('A24:F24')
ws_emp['A24'] = 'CONTRIBUTION DETAIL BY SCHEME'
ws_emp['A24'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_emp['A24'].fill = fill(TEAL)
ws_emp['A24'].alignment = Alignment(horizontal='left', vertical='center')
ws_emp.row_dimensions[24].height = 20

contrib_rows = [
    (25, 'Pension / Retirement Contribution %', f"=IFERROR(VLOOKUP(C2,{CD},7,0),0)"),
    (26, 'Pension / Retirement (USD/yr)',         f"=C16*C25"),
    (27, 'Health / Social Insurance %',           f"=IFERROR(VLOOKUP(C2,{CD},8,0),0)"),
    (28, 'Health / Social Insurance (USD/yr)',    f"=C16*C27"),
    (29, 'Other Mandatory Contributions %',       f"=IFERROR(VLOOKUP(C2,{CD},9,0),0)"),
    (30, 'Other Mandatory Contributions (USD/yr)',f"=C16*C29"),
    (31, 'Employer Benefits (health, life, etc.)',f"=C20"),
    (32, 'Govt Leave Pay Subsidy (maternity)',    f"=IFERROR(VLOOKUP(C2,{CD},12,0),0)"),
]

for row, label, expr in contrib_rows:
    lbl(ws_emp, row, 1, label, indent=2)
    ws_emp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    fmt_str = FMT_PCT if row in [25,27,29,32] else FMT_USD
    formula(ws_emp, row, 3, expr, fmt=fmt_str, green=(row in [26,28,30,31]))
    ws_emp.row_dimensions[row].height = 18

# Local currency equivalent
ws_emp.merge_cells('A34:F34')
ws_emp['A34'] = 'LOCAL CURRENCY EQUIVALENT'
ws_emp['A34'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_emp['A34'].fill = fill(NAVY)
ws_emp['A34'].alignment = Alignment(horizontal='left', vertical='center')

lbl(ws_emp, 35, 1, 'Total Employment Cost in Local Currency', indent=2)
ws_emp.merge_cells('A35:B35')
fx_formula = f"=IFERROR(C22*VLOOKUP(C2,{CD},6,0),\"-\")"
c_local = formula(ws_emp, 35, 3, fx_formula, fmt='#,##0', green=True)

lbl(ws_emp, 36, 1, 'Monthly Cost in Local Currency', indent=2)
ws_emp.merge_cells('A36:B36')
formula(ws_emp, 36, 3, "=IFERROR(C35/12,\"-\")", fmt='#,##0', green=True)

for r in [34,35,36]: ws_emp.row_dimensions[r].height = 18

set_col_widths(ws_emp, {'A':12,'B':22,'C':18,'D':15,'E':15,'F':10})

# ─────────────────────────────────────────────────────────────
# SHEET 5: REPLACEMENT COST
# ─────────────────────────────────────────────────────────────
ws_rep = wb.create_sheet('Replacement Cost')
ws_rep.sheet_properties.tabColor = 'C0392B'

ws_rep.merge_cells('A1:F1')
t = ws_rep['A1']
t.value = 'EMPLOYEE REPLACEMENT COST CALCULATOR'
t.font = Font(name='Arial', bold=True, size=12, color=WHITE)
t.fill = fill('9B2828')
t.alignment = Alignment(horizontal='center', vertical='center')
ws_rep.row_dimensions[1].height = 30

# Location selector (same approach as Employment Cost)
ws_rep.merge_cells('A2:B2')
ws_rep['A2'] = 'Select Location:'
ws_rep['A2'].font = font(bold=True)
ws_rep['A2'].alignment = Alignment(horizontal='right', vertical='center')
ws_rep['A2'].fill = fill(LTRED)
ws_rep['A2'].border = border()

loc_r = ws_rep.cell(2, 3, 'Canada')
loc_r.font = font(color=INPUTBLUE, bold=True, size=11)
loc_r.fill = fill(INPUTYELLOW)
loc_r.border = border()
loc_r.alignment = Alignment(horizontal='left', vertical='center')

dv_r = DataValidation(
    type='list',
    formula1='"Canada,Kenya,Ethiopia,Nigeria,Senegal,Tanzania,India,Bangladesh,Indonesia,Pakistan,Philippines"',
    allow_blank=False, showDropDown=False
)
ws_rep.add_data_validation(dv_r)
dv_r.add('C2')

ws_rep.merge_cells('D2:F2')
ws_rep['D2'].value = '← Salary & assumptions from Assumptions tab'
ws_rep['D2'].font = font(color='888888', size=9, italic=True)

# Background check by country (hardcoded lookup table in hidden area)
# We'll embed it at cols H-I rows 3-14
bg_checks = [
    ('Canada',220), ('Kenya',70), ('Ethiopia',55), ('Nigeria',65),
    ('Senegal',65), ('Tanzania',65), ('India',90), ('Bangladesh',70),
    ('Indonesia',80), ('Pakistan',70), ('Philippines',80)
]
ws_rep['H2'] = 'Country'
ws_rep['I2'] = 'BG Check $'
ws_rep['H2'].font = font(bold=True, size=9)
ws_rep['I2'].font = font(bold=True, size=9)
for ri, (country, cost) in enumerate(bg_checks, 3):
    ws_rep.cell(ri, 8, country)
    ws_rep.cell(ri, 9, cost)
    ws_rep.cell(ri, 8).font = font(size=9)
    ws_rep.cell(ri, 9).font = font(size=9, color=INPUTBLUE)
# Note on the hidden table
ws_rep.column_dimensions['H'].width = 14
ws_rep.column_dimensions['I'].width = 12

ASMP = 'Assumptions'
sal_ref = f'={ASMP}!B5'

# Section: Replacement Cost Breakdown
ws_rep.merge_cells('A4:F4')
ws_rep['A4'] = 'REPLACEMENT COST BREAKDOWN'
ws_rep['A4'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_rep['A4'].fill = fill('9B2828')
ws_rep['A4'].alignment = Alignment(horizontal='left', vertical='center')
ws_rep.row_dimensions[4].height = 20

daily_rate_expr = f'={ASMP}!B5/260'   # Annual ÷ 260 working days
bg_vlookup = '=IFERROR(VLOOKUP(C2,$H$3:$I$13,2,0),100)'

rep_cost_rows = [
    (5,  'Base Annual Salary (from Assumptions)',      f'={ASMP}!B5',                                    FMT_USD),
    (6,  'Daily Rate (salary / 260 working days)',      f'={ASMP}!B5/260',                                FMT_USD),
    (8,  '1. Recruitment Agency / Search Fee',          f'={ASMP}!B5*{ASMP}!B9',                         FMT_USD),
    (9,  '2. Job Advertising & Posting',                f'={ASMP}!B10',                                   FMT_USD),
    (10, '3. Interview Panel Time Cost',                f'=({ASMP}!B5/260)*{ASMP}!B11/8',                FMT_USD),
    (11, '4. Background & Reference Checks',            f'=IFERROR(VLOOKUP(C2,$H$3:$I$13,2,0),100)',     FMT_USD),
    (12, '5. Onboarding & Training',                    f'={ASMP}!B12',                                   FMT_USD),
    (13, '6. Productivity Loss During Ramp-up',         f'={ASMP}!B5*({ASMP}!B13/12)*(1-{ASMP}!B14)',   FMT_USD),
    (14, '7. Knowledge & Handover Loss',                f'={ASMP}!B5*{ASMP}!B15',                        FMT_USD),
]

for row, label, expr, fmt_str in rep_cost_rows:
    if row == 7:
        ws_rep.row_dimensions[7].height = 6
        continue
    lbl(ws_rep, row, 1, label, indent=(2 if row > 7 else 0), bold=(row in [5,6]))
    ws_rep.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    formula(ws_rep, row, 3, expr, fmt=fmt_str, green=(row in [5,8,9,10,12,13,14]))
    ws_rep.row_dimensions[row].height = 18

# Total
ws_rep.row_dimensions[15].height = 22
for col in range(1, 7):
    ws_rep.cell(15, col).fill = fill('9B2828')
    ws_rep.cell(15, col).font = Font(name='Arial', bold=True, color=WHITE, size=11)
lbl(ws_rep, 15, 1, 'TOTAL REPLACEMENT COST (USD)', bold=True)
ws_rep.merge_cells('A15:B15')
t15 = formula(ws_rep, 15, 3, '=C8+C9+C10+C11+C12+C13+C14', fmt=FMT_USD)
t15.font = Font(name='Arial', bold=True, size=12, color=WHITE)
t15.fill = fill('9B2828')

# As % of salary
ws_rep.row_dimensions[16].height = 18
lbl(ws_rep, 16, 1, '   As % of Annual Base Salary', indent=2)
ws_rep.merge_cells('A16:B16')
formula(ws_rep, 16, 3, '=IFERROR(C15/C5,0)', fmt=FMT_PCT)

# Timeline section
ws_rep.merge_cells('A18:F18')
ws_rep['A18'] = 'ESTIMATED REPLACEMENT TIMELINE'
ws_rep['A18'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_rep['A18'].fill = fill(NAVY)
ws_rep['A18'].alignment = Alignment(horizontal='left', vertical='center')
ws_rep.row_dimensions[18].height = 20

tl_rows = [
    (19, 'Vacancy / Active Recruitment Period',  '~8 weeks (assumption)',  ''),
    (20, 'Notice & Handover Period',              'Varies by country',      f"=IFERROR(VLOOKUP(C2,{CD},0,0),30)&\" days statutory\""),
    (21, 'Onboarding & Ramp-up Period',           f'Ramp-up months × 4.33 weeks', f'=ROUND({ASMP}!B13*4.33,0)&" weeks"'),
    (22, 'TOTAL: Weeks to Full Productivity',     '', f'=ROUND(8 + 8 + {ASMP}!B13*4.33, 0)&" weeks (est.)"'),
]

for row, label, note, expr in tl_rows:
    lbl(ws_rep, row, 1, label, bold=(row == 22), indent=(0 if row == 22 else 2))
    ws_rep.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    nc = ws_rep.cell(row, 3, note if not expr else '')
    nc.font = font(color='888888', size=9, italic=True)
    nc.border = border()
    ws_rep.row_dimensions[row].height = 18
    if expr:
        nc.value = expr
        nc.font = font(color=FORMULABLK)

set_col_widths(ws_rep, {'A':12,'B':32,'C':18,'D':15,'E':10,'F':10})

# ─────────────────────────────────────────────────────────────
# SHEET 6: LEAVE COST
# ─────────────────────────────────────────────────────────────
ws_lv = wb.create_sheet('Leave Cost')
ws_lv.sheet_properties.tabColor = '1A5E4E'

ws_lv.merge_cells('A1:F1')
t = ws_lv['A1']
t.value = 'STAFF LEAVE COST CALCULATOR'
t.font = Font(name='Arial', bold=True, size=12, color=WHITE)
t.fill = fill(TEAL)
t.alignment = Alignment(horizontal='center', vertical='center')
ws_lv.row_dimensions[1].height = 30

# Inputs
ws_lv.merge_cells('A2:B2')
ws_lv['A2'] = 'Select Location:'
ws_lv['A2'].font = font(bold=True)
ws_lv['A2'].alignment = Alignment(horizontal='right', vertical='center')
ws_lv['A2'].fill = fill(LTGREEN)
ws_lv['A2'].border = border()

loc_lv = ws_lv.cell(2, 3, 'Canada')
loc_lv.font = font(color=INPUTBLUE, bold=True, size=11)
loc_lv.fill = fill(INPUTYELLOW)
loc_lv.border = border()
loc_lv.alignment = Alignment(horizontal='left', vertical='center')

dv_lv = DataValidation(
    type='list',
    formula1='"Canada,Kenya,Ethiopia,Nigeria,Senegal,Tanzania,India,Bangladesh,Indonesia,Pakistan,Philippines"',
    allow_blank=False, showDropDown=False
)
ws_lv.add_data_validation(dv_lv)
dv_lv.add('C2')

ws_lv.merge_cells('A3:B3')
ws_lv['A3'] = 'Leave Type:'
ws_lv['A3'].font = font(bold=True)
ws_lv['A3'].alignment = Alignment(horizontal='right', vertical='center')
ws_lv['A3'].fill = fill(LTGREEN)
ws_lv['A3'].border = border()

lv_type = ws_lv.cell(3, 3, 'Maternity / Parental')
lv_type.font = font(color=INPUTBLUE, bold=True)
lv_type.fill = fill(INPUTYELLOW)
lv_type.border = border()
lv_type.alignment = Alignment(horizontal='left', vertical='center')

dv_lt = DataValidation(
    type='list',
    formula1='"Maternity / Parental,Medical / Sick Leave,Annual Leave,Long Service Leave"',
    showDropDown=False
)
ws_lv.add_data_validation(dv_lt)
dv_lt.add('C3')

ws_lv.merge_cells('A4:B4')
ws_lv['A4'] = 'Leave Duration (weeks):'
ws_lv['A4'].font = font(bold=True)
ws_lv['A4'].alignment = Alignment(horizontal='right', vertical='center')
ws_lv['A4'].fill = fill(LTGREEN)
ws_lv['A4'].border = border()

lv_dur = ws_lv.cell(4, 3, 17)
lv_dur.font = font(color=INPUTBLUE, bold=True)
lv_dur.fill = fill(INPUTYELLOW)
lv_dur.border = border()
lv_dur.alignment = Alignment(horizontal='right', vertical='center')
lv_dur.number_format = '0'

ws_lv.merge_cells('A5:B5')
ws_lv['A5'] = 'Backfill Required?'
ws_lv['A5'].font = font(bold=True)
ws_lv['A5'].alignment = Alignment(horizontal='right', vertical='center')
ws_lv['A5'].fill = fill(LTGREEN)
ws_lv['A5'].border = border()

lv_bf = ws_lv.cell(5, 3, 'Full backfill (hire temp)')
lv_bf.font = font(color=INPUTBLUE)
lv_bf.fill = fill(INPUTYELLOW)
lv_bf.border = border()
lv_bf.alignment = Alignment(horizontal='left', vertical='center')

dv_bf = DataValidation(
    type='list',
    formula1='"Full backfill (hire temp),Partial redistribution (40%),No backfill"',
    showDropDown=False
)
ws_lv.add_data_validation(dv_bf)
dv_bf.add('C5')

for r in [2,3,4,5]: ws_lv.row_dimensions[r].height = 20

# Cost breakdown
ws_lv.merge_cells('A7:F7')
ws_lv['A7'] = 'LEAVE COST BREAKDOWN'
ws_lv['A7'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_lv['A7'].fill = fill(TEAL)
ws_lv['A7'].alignment = Alignment(horizontal='left', vertical='center')
ws_lv.row_dimensions[7].height = 20

# Weekly salary — use direct cell-reference-friendly formulas (avoid double-= expansion)
weekly_sal    = f'={ASMP}!B5/52'
gross_leave   = f'=({ASMP}!B5/52)*C4'
# govt offset stored as a NEGATIVE value so C12 = C10+C11
govt_offset_neg = f"=-IFERROR(({ASMP}!B5/52)*C4*VLOOKUP(C2,{CD},12,0)/100,0)"
# net salary = gross (C10) + offset (C11 is already negative)
net_sal       = '=C10+C11'
# contributions on net salary paid (references C12 directly)
mand_on_leave = f"=IFERROR(C12*VLOOKUP(C2,{CD},10,0)/100,0)"
backfill_expr = f'=IF(C5="Full backfill (hire temp)",{ASMP}!B18*C4*5,IF(C5="Partial redistribution (40%)",{ASMP}!B18*C4*5*0.4,0))'
prod_impact   = f'=({ASMP}!B5/52)*C4*{ASMP}!B22'

lv_breakdown = [
    (8,  'Annual Base Salary (from Assumptions)',       f'={ASMP}!B5',   FMT_USD),
    (9,  'Weekly Salary',                               weekly_sal,      FMT_USD),
    (10, 'Gross Salary During Leave',                   gross_leave,     FMT_USD),
    (11, 'Less: Government / Statutory Leave Pay',      govt_offset_neg, FMT_USD),
    (12, 'Net Salary Cost to Employer',                 net_sal,         FMT_USD),
    (13, 'Employer Contributions on Salary Paid',       mand_on_leave,   FMT_USD),
    (14, 'Backfill / Temporary Cover Cost',             backfill_expr,   FMT_USD),
    (15, 'Team Productivity Impact (est.)',             prod_impact,     FMT_USD),
    (16, 'Admin & Transition Overhead',                 '=450',          FMT_USD),
]

for row, label, expr, fmt_str in lv_breakdown:
    lbl(ws_lv, row, 1, label, indent=(2 if row > 9 else 0), bold=(row in [8,9]))
    ws_lv.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    formula(ws_lv, row, 3, expr, fmt=fmt_str, green=(row in [8,10]))
    ws_lv.row_dimensions[row].height = 18

# Total
ws_lv.row_dimensions[17].height = 22
for col in range(1, 7):
    ws_lv.cell(17, col).fill = fill(TEAL)
    ws_lv.cell(17, col).font = Font(name='Arial', bold=True, color=WHITE, size=11)
lbl(ws_lv, 17, 1, 'TOTAL LEAVE COST TO ORGANISATION (USD)', bold=True)
ws_lv.merge_cells('A17:B17')
t17 = formula(ws_lv, 17, 3, '=SUM(C12:C16)', fmt=FMT_USD)
t17.font = Font(name='Arial', bold=True, size=12, color=WHITE)
t17.fill = fill(TEAL)

# Leave entitlements section
ws_lv.merge_cells('A19:F19')
ws_lv['A19'] = 'STATUTORY LEAVE ENTITLEMENTS — SELECTED COUNTRY'
ws_lv['A19'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
ws_lv['A19'].fill = fill(NAVY)
ws_lv['A19'].alignment = Alignment(horizontal='left', vertical='center')
ws_lv.row_dimensions[19].height = 20

ent_rows = [
    (20, 'Maternity Leave (statutory)',  f"=IFERROR(VLOOKUP(C2,{CD},13,0),\"—\")&\" weeks\""),
    (21, 'Annual Leave (statutory)',     f"=IFERROR(VLOOKUP(C2,{CD},14,0),\"—\")&\" working days/yr\""),
    (22, 'Govt Leave Pay Subsidy',       f"=IFERROR(IF(VLOOKUP(C2,{CD},12,0)=0,\"None — employer bears full cost\",VLOOKUP(C2,{CD},12,0)&\"% of salary covered by govt\"),\"—\")"),
    (23, 'Total Mandatory Employer Rate',f"=IFERROR(VLOOKUP(C2,{CD},10,0),0)", FMT_PCT),
]

for item in ent_rows:
    row = item[0]; label = item[1]; expr = item[2]
    fmt_str = item[3] if len(item) > 3 else '@'
    lbl(ws_lv, row, 1, label, indent=2)
    ws_lv.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = formula(ws_lv, row, 3, expr, green=True)
    if fmt_str != '@': c.number_format = fmt_str
    ws_lv.row_dimensions[row].height = 18

set_col_widths(ws_lv, {'A':12,'B':34,'C':22,'D':15,'E':10,'F':10})

# ─────────────────────────────────────────────────────────────
# SHEET 7: COMPARE LOCATIONS
# ─────────────────────────────────────────────────────────────
ws_cmp = wb.create_sheet('Compare Locations')
ws_cmp.sheet_properties.tabColor = '4DC8A0'

ws_cmp.merge_cells('A1:I1')
t = ws_cmp['A1']
t.value = 'LOCATION COST COMPARISON — SIDE BY SIDE (UP TO 4 LOCATIONS)'
t.font = Font(name='Arial', bold=True, size=12, color=WHITE)
t.fill = fill(NAVY)
t.alignment = Alignment(horizontal='center', vertical='center')
ws_cmp.row_dimensions[1].height = 30

# Location selectors
ws_cmp.merge_cells('A2:B2')
ws_cmp['A2'] = 'Base Salary (USD):'
ws_cmp['A2'].font = font(bold=True)
ws_cmp['A2'].alignment = Alignment(horizontal='right', vertical='center')
ws_cmp['A2'].border = border()

sal_cmp = ws_cmp.cell(2, 3, f'={ASMP}!B5')
sal_cmp.font = font(color=LINKGREEN, bold=True)
sal_cmp.border = border()
sal_cmp.number_format = FMT_USD
ws_cmp.merge_cells('D2:I2')
ws_cmp['D2'].value = '← Pulling from Assumptions tab. Change salary there to update all columns.'
ws_cmp['D2'].font = font(color='888888', size=9, italic=True)

ws_cmp.row_dimensions[2].height = 20

# Headers row
hdr(ws_cmp, 3, 1, 'Cost Component', bg=NAVY, span=2)
for ci, default in enumerate(['Canada', 'Kenya', 'India', 'Nigeria'], 3):
    c = ws_cmp.cell(4, ci, default)
    c.font = font(color=INPUTBLUE, bold=True, size=11)
    c.fill = fill(INPUTYELLOW)
    c.border = border()
    c.alignment = Alignment(horizontal='center', vertical='center')

    dv_cmp = DataValidation(
        type='list',
        formula1='"Canada,Kenya,Ethiopia,Nigeria,Senegal,Tanzania,India,Bangladesh,Indonesia,Pakistan,Philippines"',
        showDropDown=False
    )
    ws_cmp.add_data_validation(dv_cmp)
    dv_cmp.add(ws_cmp.cell(4, ci))

ws_cmp.row_dimensions[4].height = 22

# Column letter helper for data columns
def cc(col_idx): return get_column_letter(col_idx)

# Build comparison rows
# Each data column (3,4,5,6) references: C4/D4/E4/F4 = selected country
# We need VLOOKUP per column

def cmp_formula(country_cell, col_idx):
    """Return VLOOKUP expression for given country cell reference."""
    return country_cell  # Just the cell ref — formulas built inline below

data_cols = [3, 4, 5, 6]  # columns C, D, E, F

def cmp_row(ws, row, label, formulas, fmt_str=None, is_sub=False, is_grand=False, bold=False):
    ws.row_dimensions[row].height = 20
    # Label (merged A:B)
    lc = ws.cell(row, 1, label)
    lc.font = font(bold=(bold or is_sub or is_grand), color=(WHITE if is_grand else FORMULABLK))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    lc.border = border()
    lc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if is_grand: lc.fill = fill(NAVY)
    elif is_sub: lc.fill = fill(MIDGRAY)
    else: lc.fill = fill(WHITE)

    for ci, expr in zip(data_cols, formulas):
        c = ws.cell(row, ci, expr)
        c.font = font(bold=(is_sub or is_grand), color=(WHITE if is_grand else (LINKGREEN if not is_sub else FORMULABLK)))
        c.border = border()
        c.alignment = Alignment(horizontal='right', vertical='center')
        if fmt_str: c.number_format = fmt_str
        if is_grand: c.fill = fill(NAVY)
        elif is_sub: c.fill = fill(MIDGRAY)

def mk_vlookup(row_ltr, col_idx, lookup_col):
    """VLOOKUP into Country Data for a given comparison column."""
    country_ref = f'{cc(col_idx)}4'
    return f"=IFERROR(VLOOKUP({country_ref},{CD},{lookup_col},0),0)"

def mk_pct_vlookup(row_ltr, col_idx, lookup_col):
    return f"=IFERROR(VLOOKUP({cc(col_idx)}4,{CD},{lookup_col},0),0)"

# Row definitions
sal_expr = [f"=C$2"] * 4
sal_expr = [f"={ASMP}!$B$5"] * 4

r = 5
# Section header
ws_cmp.merge_cells(f'A{r}:I{r}')
sh = ws_cmp.cell(r, 1, 'SECTION 1 — ANNUAL EMPLOYMENT COST')
sh.font = Font(name='Arial', bold=True, size=10, color=WHITE)
sh.fill = fill(TEAL)
sh.alignment = Alignment(horizontal='left', vertical='center')
ws_cmp.row_dimensions[r].height = 20
r += 1

cmp_row(ws_cmp, r, 'Base Annual Salary (USD)',
        [f'={ASMP}!$B$5'] * 4, FMT_USD)
r += 1

cmp_row(ws_cmp, r, 'Mandatory Employer Contributions (%)',
        [mk_vlookup(r, c, 10) for c in data_cols], FMT_PCT)
r += 1

mand_row = r
cmp_row(ws_cmp, r, 'Mandatory Contributions (USD)',
        [f'={ASMP}!$B$5*IFERROR(VLOOKUP({cc(c)}$4,{CD},10,0),0)' for c in data_cols], FMT_USD)
r += 1

ben_row = r
cmp_row(ws_cmp, r, 'Employer Benefits — Health, Life (USD)',
        [f'={ASMP}!$B$5*IFERROR(VLOOKUP({cc(c)}$4,{CD},11,0),0)' for c in data_cols], FMT_USD)
r += 1

emp_total_row = r
cmp_row(ws_cmp, r, 'TOTAL ANNUAL EMPLOYMENT COST (USD)',
        [f'={ASMP}!$B$5+{ASMP}!$B$5*IFERROR(VLOOKUP({cc(c)}$4,{CD},10,0),0)+{ASMP}!$B$5*IFERROR(VLOOKUP({cc(c)}$4,{CD},11,0),0)'
         for c in data_cols], FMT_USD, is_sub=True)
r += 1

cmp_row(ws_cmp, r, '  Employer Burden % (above base salary)',
        [f'=IFERROR(({cc(c)}{emp_total_row}-{ASMP}!$B$5)/{ASMP}!$B$5,0)' for c in data_cols], FMT_PCT)
r += 2

# Section 2: Replacement
ws_cmp.merge_cells(f'A{r}:I{r}')
sh2 = ws_cmp.cell(r, 1, 'SECTION 2 — REPLACEMENT COST (if employee departs)')
sh2.font = Font(name='Arial', bold=True, size=10, color=WHITE)
sh2.fill = fill('9B2828')
sh2.alignment = Alignment(horizontal='left', vertical='center')
ws_cmp.row_dimensions[r].height = 20
r += 1

cmp_row(ws_cmp, r, 'Recruitment (agency + ads + interviews)',
        [f'={ASMP}!$B$5*{ASMP}!$B$9+{ASMP}!$B$10+({ASMP}!$B$5/260)*{ASMP}!$B$11/8'] * 4, FMT_USD)
r += 1

bg_checks_inline = {
    'Canada': 220, 'Kenya': 70, 'Ethiopia': 55, 'Nigeria': 65,
    'Senegal': 65, 'Tanzania': 65, 'India': 90, 'Bangladesh': 70,
    'Indonesia': 80, 'Pakistan': 70, 'Philippines': 80
}
# embed bg check VLOOKUP pointing to our helper table
cmp_row(ws_cmp, r, 'Background Checks (country-adjusted)',
        [f'=IFERROR(VLOOKUP({cc(c)}$4,Replacement!$H$3:$I$13,2,0),100)' for c in data_cols], FMT_USD)
r += 1

cmp_row(ws_cmp, r, 'Onboarding & Training',
        [f'={ASMP}!$B$12'] * 4, FMT_USD)
r += 1

cmp_row(ws_cmp, r, 'Ramp-up Productivity Loss',
        [f'={ASMP}!$B$5*({ASMP}!$B$13/12)*(1-{ASMP}!$B$14)'] * 4, FMT_USD)
r += 1

cmp_row(ws_cmp, r, 'Knowledge & Handover Loss',
        [f'={ASMP}!$B$5*{ASMP}!$B$15'] * 4, FMT_USD)
r += 1

rep_total_row = r
cmp_row(ws_cmp, r, 'TOTAL REPLACEMENT COST (USD)',
        [f'={ASMP}!$B$5*{ASMP}!$B$9+{ASMP}!$B$10+({ASMP}!$B$5/260)*{ASMP}!$B$11/8+'
         f'IFERROR(VLOOKUP({cc(c)}$4,Replacement!$H$3:$I$13,2,0),100)+'
         f'{ASMP}!$B$12+{ASMP}!$B$5*({ASMP}!$B$13/12)*(1-{ASMP}!$B$14)+{ASMP}!$B$5*{ASMP}!$B$15'
         for c in data_cols], FMT_USD, is_sub=True)
r += 1

cmp_row(ws_cmp, r, '  Replacement as % of Annual Salary',
        [f'=IFERROR({cc(c)}{rep_total_row}/{ASMP}!$B$5,0)' for c in data_cols], FMT_PCT)
r += 2

# Section 3: Leave
ws_cmp.merge_cells(f'A{r}:I{r}')
sh3 = ws_cmp.cell(r, 1, 'SECTION 3 — LEAVE COST (17 weeks maternity, full backfill)')
sh3.font = Font(name='Arial', bold=True, size=10, color=WHITE)
sh3.fill = fill(TEAL)
sh3.alignment = Alignment(horizontal='left', vertical='center')
ws_cmp.row_dimensions[r].height = 20
r += 1

cmp_row(ws_cmp, r, 'Statutory Maternity Leave (weeks)',
        [f'=IFERROR(VLOOKUP({cc(c)}$4,{CD},13,0),0)' for c in data_cols], '0')
r += 1

cmp_row(ws_cmp, r, 'Govt Leave Pay Subsidy (%)',
        [f'=IFERROR(VLOOKUP({cc(c)}$4,{CD},12,0),0)' for c in data_cols], FMT_PCT)
r += 1

# 17 wk maternity leave cost
cmp_row(ws_cmp, r, 'Net Salary Cost to Employer (17 wks)',
        [f'=({ASMP}!$B$5/52)*17*(1-IFERROR(VLOOKUP({cc(c)}$4,{CD},12,0),0)/100)' for c in data_cols], FMT_USD)
r += 1

cmp_row(ws_cmp, r, 'Full Backfill Cost (temp @ daily rate)',
        [f'={ASMP}!$B$18*17*5'] * 4, FMT_USD)
r += 1

lv_total_row = r
cmp_row(ws_cmp, r, 'TOTAL LEAVE COST — 17 wks (USD)',
        [f'=({ASMP}!$B$5/52)*17*(1-IFERROR(VLOOKUP({cc(c)}$4,{CD},12,0),0)/100)+'
         f'{ASMP}!$B$18*17*5+({ASMP}!$B$5/52)*17*{ASMP}!$B$22+450'
         for c in data_cols], FMT_USD, is_sub=True)
r += 2

# Grand total
cmp_row(ws_cmp, r, 'WORST CASE: Employment + Replacement Cost',
        [f'=IFERROR({cc(c)}{emp_total_row}+{cc(c)}{rep_total_row},0)' for c in data_cols],
        FMT_USD, is_grand=True)
r += 1

cmp_row(ws_cmp, r, '  Additional if Leave also occurs (+17 wks)',
        [f'=IFERROR({cc(c)}{emp_total_row}+{cc(c)}{rep_total_row}+{cc(c)}{lv_total_row},0)'
         for c in data_cols], FMT_USD)
r += 2

# Note
ws_cmp.merge_cells(f'A{r}:I{r}')
note = ws_cmp.cell(r, 1, 'Notes: Salary pulled from Assumptions tab. Leave cost uses 17 weeks + full backfill. Recruitment uses current assumptions. Select countries using dropdown in row 4.')
note.font = font(color='888888', size=9, italic=True)
note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_cmp.row_dimensions[r].height = 28

set_col_widths(ws_cmp, {'A':4,'B':34,'C':18,'D':18,'E':18,'F':18,'G':5,'H':5,'I':5})

# ─────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print('SAVED:', OUTPUT)
